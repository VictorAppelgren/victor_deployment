#!/usr/bin/env python3
"""
SagaLabs Seed Round Budget Generator - WITH EXCEL FORMULAS

Generates Excel file with:
- INTERACTIVE: All calculations use Excel formulas, not hardcoded values
- Change assumptions and see results update automatically
- 24-month cashflow projection
- Team build timeline
- Multiple scenarios

Usage:
    python generate_budget.py [--output path/to/output.xlsx]
"""

import argparse
from datetime import datetime
from pathlib import Path

# Try to import required libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install openpyxl")
    exit(1)

# =============================================================================
# DEFAULT VALUES - These will be put in editable cells in Excel
# =============================================================================

# Raise
DEFAULT_RAISE_AMOUNT = 35_000_000
DEFAULT_DILUTION = 0.20

# Salaries - HIGHER as requested
DEFAULT_AVG_BASE_SALARY = 90_000  # SEK/month - increased from 77K
DEFAULT_ARBETSGIVARAVGIFT = 0.32  # 32% - increased from 30%

# Pricing
DEFAULT_MONTHLY_LICENSE_FEE = 100_000

# Costs - INCREASED significantly
DEFAULT_COMPUTE_BASE = 50_000  # Increased from 20K
DEFAULT_COMPUTE_PER_CUSTOMER = 12_000  # Increased from 8K
DEFAULT_DATA_BASE = 80_000  # Increased from 20K - data is expensive!
DEFAULT_DATA_PER_CUSTOMER = 5_000  # Increased from 2K
DEFAULT_INFRASTRUCTURE_BASE = 25_000  # Increased from 10K
DEFAULT_INFRASTRUCTURE_PER_EMPLOYEE = 3_000  # Increased from 2K
DEFAULT_OFFICE_PER_EMPLOYEE = 6_000  # Increased from 5K
DEFAULT_SALES_MARKETING_BASE = 30_000  # Increased from 15K
DEFAULT_SALES_MARKETING_PER_CUSTOMER = 5_000  # Increased from 3K
DEFAULT_ADMIN_LEGAL_BASE = 25_000  # Increased from 15K
DEFAULT_BUFFER_PERCENT = 0.10  # Increased from 5% to 10%

# Customer growth - DELAYED: First paying customer M5 (not M3)
# Slower ramp: 0,0,0,0,1,1,1,2,2,2,3,3 = 15 customers Y1
DEFAULT_NEW_CUSTOMERS = [0, 0, 0, 0, 1, 1, 1, 2, 2, 2, 3, 3, 3, 4, 4, 4, 5, 5, 5, 6, 6, 6, 7, 7]

# Team growth - aggressive to use the capital
DEFAULT_TEAM_SIZE = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 17, 19, 21, 23, 25, 26, 27, 28, 29, 30, 30, 30]


def create_assumptions_sheet(wb: Workbook):
    """Create the Assumptions sheet with all editable parameters."""
    ws = wb.active
    ws.title = "Assumptions"

    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, size=12)
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid")  # Yellow for editable

    # Title
    ws["A1"] = "SAGALABS FINANCIAL MODEL - ASSUMPTIONS"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:D1")

    ws["A2"] = "Yellow cells are editable - change them to see impact on projections"
    ws["A2"].font = Font(italic=True, size=10)

    row = 4

    # === RAISE SECTION ===
    ws.cell(row=row, column=1, value="RAISE DETAILS").font = header_font
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    ws.cell(row=row, column=1, value="Raise Amount (SEK)")
    ws.cell(row=row, column=2, value=DEFAULT_RAISE_AMOUNT)
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=2).fill = input_fill
    ws["B5"].font = Font(bold=True)  # Named reference: raise_amount
    row += 1

    ws.cell(row=row, column=1, value="Dilution (%)")
    ws.cell(row=row, column=2, value=DEFAULT_DILUTION)
    ws.cell(row=row, column=2).number_format = '0%'
    ws.cell(row=row, column=2).fill = input_fill
    row += 1

    # Calculated
    ws.cell(row=row, column=1, value="Pre-Money Valuation")
    ws.cell(row=row, column=2, value="=B5/B6-B5")
    ws.cell(row=row, column=2).number_format = '#,##0'
    row += 1

    ws.cell(row=row, column=1, value="Post-Money Valuation")
    ws.cell(row=row, column=2, value="=B5/B6")
    ws.cell(row=row, column=2).number_format = '#,##0'
    row += 2

    # === SALARY SECTION ===
    ws.cell(row=row, column=1, value="SALARY COSTS (SWEDEN)").font = header_font
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    salary_row = row  # Remember for formulas
    ws.cell(row=row, column=1, value="Avg Base Salary (SEK/month)")
    ws.cell(row=row, column=2, value=DEFAULT_AVG_BASE_SALARY)
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=2).fill = input_fill
    row += 1

    avgift_row = row
    ws.cell(row=row, column=1, value="Arbetsgivaravgift (%)")
    ws.cell(row=row, column=2, value=DEFAULT_ARBETSGIVARAVGIFT)
    ws.cell(row=row, column=2).number_format = '0%'
    ws.cell(row=row, column=2).fill = input_fill
    row += 1

    total_cost_row = row
    ws.cell(row=row, column=1, value="Total Cost per Employee")
    ws.cell(row=row, column=2, value=f"=B{salary_row}*(1+B{avgift_row})")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3, value="= Base * (1 + Avgift)")
    ws.cell(row=row, column=3).font = Font(italic=True, color="666666")
    row += 2

    # === PRICING SECTION ===
    ws.cell(row=row, column=1, value="PRICING").font = header_font
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    license_row = row
    ws.cell(row=row, column=1, value="Monthly License Fee (SEK)")
    ws.cell(row=row, column=2, value=DEFAULT_MONTHLY_LICENSE_FEE)
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=2).fill = input_fill
    row += 1

    ws.cell(row=row, column=1, value="Annual Contract Value")
    ws.cell(row=row, column=2, value=f"=B{license_row}*12")
    ws.cell(row=row, column=2).number_format = '#,##0'
    row += 2

    # === COST SECTION ===
    ws.cell(row=row, column=1, value="OPERATING COSTS").font = header_font
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    # Store row numbers for formula references
    cost_rows = {}

    costs = [
        ("Compute - Base (SEK/month)", DEFAULT_COMPUTE_BASE, "compute_base"),
        ("Compute - Per Customer", DEFAULT_COMPUTE_PER_CUSTOMER, "compute_per_cust"),
        ("Data - Base (SEK/month)", DEFAULT_DATA_BASE, "data_base"),
        ("Data - Per Customer", DEFAULT_DATA_PER_CUSTOMER, "data_per_cust"),
        ("Infrastructure - Base", DEFAULT_INFRASTRUCTURE_BASE, "infra_base"),
        ("Infrastructure - Per Employee", DEFAULT_INFRASTRUCTURE_PER_EMPLOYEE, "infra_per_emp"),
        ("Office - Per Employee", DEFAULT_OFFICE_PER_EMPLOYEE, "office_per_emp"),
        ("Sales/Marketing - Base", DEFAULT_SALES_MARKETING_BASE, "sales_base"),
        ("Sales/Marketing - Per Customer", DEFAULT_SALES_MARKETING_PER_CUSTOMER, "sales_per_cust"),
        ("Admin & Legal - Base", DEFAULT_ADMIN_LEGAL_BASE, "admin_base"),
        ("Buffer (%)", DEFAULT_BUFFER_PERCENT, "buffer_pct"),
    ]

    for label, value, key in costs:
        cost_rows[key] = row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if "%" in label:
            ws.cell(row=row, column=2).number_format = '0%'
        else:
            ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = input_fill
        row += 1

    row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15

    # Return important row numbers for other sheets to reference
    return {
        "raise_amount": 5,
        "dilution": 6,
        "avg_base_salary": salary_row,
        "arbetsgivaravgift": avgift_row,
        "total_cost_per_emp": total_cost_row,
        "license_fee": license_row,
        **cost_rows
    }


def create_inputs_sheet(wb: Workbook, refs: dict):
    """Create the Inputs sheet with monthly team size and customer growth."""
    ws = wb.create_sheet(title="Inputs")

    # Styles
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    input_fill = PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid")

    ws["A1"] = "MONTHLY INPUTS - Edit these to change projections"
    ws["A1"].font = title_font
    ws.merge_cells("A1:Z1")

    # Headers
    ws.cell(row=3, column=1, value="Month").font = header_font
    ws.cell(row=3, column=1).fill = header_fill

    for m in range(1, 25):
        cell = ws.cell(row=3, column=m+1, value=f"M{m}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # New Customers row (editable)
    ws.cell(row=4, column=1, value="New Customers")
    ws.cell(row=4, column=1).font = Font(bold=True)
    for m in range(24):
        cell = ws.cell(row=4, column=m+2, value=DEFAULT_NEW_CUSTOMERS[m])
        cell.fill = input_fill
        cell.alignment = Alignment(horizontal='center')

    # Total Customers (formula: cumulative sum)
    ws.cell(row=5, column=1, value="Total Customers")
    ws.cell(row=5, column=1).font = Font(bold=True)
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        if m == 0:
            ws.cell(row=5, column=col, value=f"={col_letter}4")
        else:
            prev_col = get_column_letter(col - 1)
            ws.cell(row=5, column=col, value=f"={prev_col}5+{col_letter}4")
        ws.cell(row=5, column=col).alignment = Alignment(horizontal='center')

    # Team Size row (editable)
    ws.cell(row=7, column=1, value="Team Size")
    ws.cell(row=7, column=1).font = Font(bold=True)
    for m in range(24):
        cell = ws.cell(row=7, column=m+2, value=DEFAULT_TEAM_SIZE[m])
        cell.fill = input_fill
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions["A"].width = 18
    for col in range(2, 26):
        ws.column_dimensions[get_column_letter(col)].width = 8

    return {
        "new_customers_row": 4,
        "total_customers_row": 5,
        "team_size_row": 7,
    }


def create_cashflow_sheet(wb: Workbook, assumption_refs: dict, input_refs: dict):
    """Create the Cashflow sheet with ALL FORMULAS linking to Assumptions and Inputs."""
    ws = wb.create_sheet(title="Cashflow")

    # Styles
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    section_font = Font(bold=True, size=10)

    ws["A1"] = "24-MONTH CASH FLOW - All values are formulas linked to Assumptions & Inputs"
    ws["A1"].font = title_font
    ws.merge_cells("A1:Z1")

    # Column headers
    ws.cell(row=3, column=1, value="Category").font = header_font
    ws.cell(row=3, column=1).fill = header_fill

    for m in range(1, 25):
        cell = ws.cell(row=3, column=m+1, value=f"M{m}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Total column
    cell = ws.cell(row=3, column=26, value="TOTAL/END")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

    row = 4

    # === CUSTOMERS SECTION ===
    ws.cell(row=row, column=1, value="CUSTOMERS").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    # New Customers - link to Inputs
    ws.cell(row=row, column=1, value="New Customers")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=Inputs!{col_letter}4")
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    # Total
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    new_cust_row = row
    row += 1

    # Total Customers - link to Inputs
    ws.cell(row=row, column=1, value="Total Customers")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=Inputs!{col_letter}5")
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    # End value
    ws.cell(row=row, column=26, value=f"=Y{row}")
    total_cust_row = row
    row += 2

    # === REVENUE SECTION ===
    ws.cell(row=row, column=1, value="REVENUE").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    # Monthly Revenue = Total Customers * License Fee
    ws.cell(row=row, column=1, value="Monthly Revenue")
    license_ref = f"Assumptions!$B${assumption_refs['license_fee']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={col_letter}{total_cust_row}*{license_ref}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    revenue_row = row
    row += 1

    # ARR (Run Rate) = Monthly Revenue * 12
    ws.cell(row=row, column=1, value="ARR (Run Rate)")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={col_letter}{revenue_row}*12")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=Y{row}")
    ws.cell(row=row, column=26).number_format = '#,##0'
    row += 2

    # === TEAM SECTION ===
    ws.cell(row=row, column=1, value="TEAM").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    # Team Size - link to Inputs
    ws.cell(row=row, column=1, value="Team Size")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=Inputs!{col_letter}7")
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=Y{row}")
    team_row = row
    row += 2

    # === EXPENSES SECTION ===
    ws.cell(row=row, column=1, value="EXPENSES").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    # Salaries = Team Size * Total Cost Per Employee
    ws.cell(row=row, column=1, value="Salaries (incl. avgift)")
    total_cost_ref = f"Assumptions!$B${assumption_refs['total_cost_per_emp']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={col_letter}{team_row}*{total_cost_ref}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    salaries_row = row
    row += 1

    # Compute = Base + (Customers * Per Customer)
    ws.cell(row=row, column=1, value="Compute")
    compute_base_ref = f"Assumptions!$B${assumption_refs['compute_base']}"
    compute_per_ref = f"Assumptions!$B${assumption_refs['compute_per_cust']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={compute_base_ref}+({col_letter}{total_cust_row}*{compute_per_ref})")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    compute_row = row
    row += 1

    # Data = Base + (Customers * Per Customer)
    ws.cell(row=row, column=1, value="Data")
    data_base_ref = f"Assumptions!$B${assumption_refs['data_base']}"
    data_per_ref = f"Assumptions!$B${assumption_refs['data_per_cust']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={data_base_ref}+({col_letter}{total_cust_row}*{data_per_ref})")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    data_row = row
    row += 1

    # Infrastructure = Base + (Team * Per Employee)
    ws.cell(row=row, column=1, value="Infrastructure")
    infra_base_ref = f"Assumptions!$B${assumption_refs['infra_base']}"
    infra_per_ref = f"Assumptions!$B${assumption_refs['infra_per_emp']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={infra_base_ref}+({col_letter}{team_row}*{infra_per_ref})")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    infra_row = row
    row += 1

    # Office = Team * Per Employee
    ws.cell(row=row, column=1, value="Office")
    office_ref = f"Assumptions!$B${assumption_refs['office_per_emp']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={col_letter}{team_row}*{office_ref}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    office_row = row
    row += 1

    # Sales & Marketing = Base + (Customers * Per Customer)
    ws.cell(row=row, column=1, value="Sales & Marketing")
    sales_base_ref = f"Assumptions!$B${assumption_refs['sales_base']}"
    sales_per_ref = f"Assumptions!$B${assumption_refs['sales_per_cust']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={sales_base_ref}+({col_letter}{total_cust_row}*{sales_per_ref})")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    sales_row = row
    row += 1

    # Admin & Legal = Base
    ws.cell(row=row, column=1, value="Admin & Legal")
    admin_ref = f"Assumptions!$B${assumption_refs['admin_base']}"
    for m in range(24):
        col = m + 2
        ws.cell(row=row, column=col, value=f"={admin_ref}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    admin_row = row
    row += 1

    # Subtotal
    ws.cell(row=row, column=1, value="Subtotal Expenses")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({col_letter}{salaries_row}:{col_letter}{admin_row})")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    subtotal_row = row
    row += 1

    # Buffer = Subtotal * Buffer %
    ws.cell(row=row, column=1, value="Buffer")
    buffer_ref = f"Assumptions!$B${assumption_refs['buffer_pct']}"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={col_letter}{subtotal_row}*{buffer_ref}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    buffer_row = row
    row += 1

    # Total Expenses
    ws.cell(row=row, column=1, value="TOTAL EXPENSES")
    ws.cell(row=row, column=1).font = Font(bold=True)
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={col_letter}{subtotal_row}+{col_letter}{buffer_row}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    ws.cell(row=row, column=26).font = Font(bold=True)
    total_exp_row = row
    row += 2

    # === CASH FLOW SECTION ===
    ws.cell(row=row, column=1, value="CASH FLOW").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    # Net Cashflow = Revenue - Total Expenses
    ws.cell(row=row, column=1, value="Net Cashflow")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={col_letter}{revenue_row}-{col_letter}{total_exp_row}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    net_cf_row = row
    row += 1

    # Cash Balance = Previous Balance + Net Cashflow (starting with Raise Amount)
    ws.cell(row=row, column=1, value="Cash Balance")
    ws.cell(row=row, column=1).font = Font(bold=True)
    raise_ref = f"Assumptions!$B$5"
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        if m == 0:
            # First month: Raise Amount + Net Cashflow
            ws.cell(row=row, column=col, value=f"={raise_ref}+{col_letter}{net_cf_row}")
        else:
            # Subsequent months: Previous Balance + Net Cashflow
            prev_col = get_column_letter(col - 1)
            ws.cell(row=row, column=col, value=f"={prev_col}{row}+{col_letter}{net_cf_row}")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=Y{row}")
    ws.cell(row=row, column=26).number_format = '#,##0'
    ws.cell(row=row, column=26).font = Font(bold=True)
    cash_row = row
    row += 2

    # === KEY METRICS ===
    ws.cell(row=row, column=1, value="KEY METRICS").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1

    # Runway (months)
    ws.cell(row=row, column=1, value="Runway (months)")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=IF({col_letter}{total_exp_row}>0,{col_letter}{cash_row}/{col_letter}{total_exp_row},999)")
        ws.cell(row=row, column=col).number_format = '0.0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    row += 1

    # Monthly Burn (when losing money)
    ws.cell(row=row, column=1, value="Monthly Burn")
    for m in range(24):
        col = m + 2
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=IF({col_letter}{net_cf_row}<0,-{col_letter}{net_cf_row},0)")
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).number_format = '#,##0'
    row += 1

    # Lowest Cash Point
    row += 1
    ws.cell(row=row, column=1, value="LOWEST CASH POINT:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=MIN(B{cash_row}:Y{cash_row})")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
    row += 1

    # Capital Used (Raise - Lowest)
    ws.cell(row=row, column=1, value="MAX CAPITAL USED:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"={raise_ref}-MIN(B{cash_row}:Y{cash_row})")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col in range(2, 27):
        ws.column_dimensions[get_column_letter(col)].width = 10


def generate_budget(output_path: Path):
    """Generate the complete budget Excel file with formulas."""
    print(f"Generating budget with formulas: {output_path}")

    # Create workbook
    wb = Workbook()

    # Create sheets in order
    assumption_refs = create_assumptions_sheet(wb)
    input_refs = create_inputs_sheet(wb, assumption_refs)
    create_cashflow_sheet(wb, assumption_refs, input_refs)

    # Save
    wb.save(output_path)
    print(f"Done! Excel file: {output_path}")
    print()
    print("HOW TO USE:")
    print("  1. Open the Excel file")
    print("  2. Go to 'Assumptions' sheet - edit yellow cells")
    print("  3. Go to 'Inputs' sheet - edit monthly team size and new customers")
    print("  4. Go to 'Cashflow' sheet - see all calculations update automatically!")
    print()
    print("Key cells to tune:")
    print("  - Assumptions!B5: Raise amount")
    print("  - Assumptions!B11: Avg base salary")
    print("  - Assumptions!B12: Arbetsgivaravgift %")
    print("  - Assumptions!B16-B26: All cost assumptions")
    print("  - Inputs!B4:Y4: New customers per month")
    print("  - Inputs!B7:Y7: Team size per month")

    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate SagaLabs budget Excel file")
    parser.add_argument("--output", "-o", help="Output Excel path")
    args = parser.parse_args()

    # Default output path
    if args.output:
        output_path = Path(args.output)
    else:
        script_dir = Path(__file__).parent
        output_dir = script_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = output_dir / f"saga_budget_{timestamp}.xlsx"

    generate_budget(output_path)


if __name__ == "__main__":
    main()
